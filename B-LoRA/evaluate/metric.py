import os
import numpy as np
import csv
from PIL import Image
import torch
import torch.nn as nn
from torchvision import models, transforms
from scipy.linalg import sqrtm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
from openpyxl.utils import get_column_letter

class StyleTransferEvaluator:
    def __init__(self, root_dir, style_dir, content_dir, methods):
        self.root_dir = root_dir
        self.style_dir = style_dir
        self.content_dir = content_dir
        self.methods = methods
        
        self.model = models.vgg19(weights='VGG19_Weights.DEFAULT').features.eval()
        # self.style_layers = ['0', '5', '10', '17', '24']
        self.style_layers = ['0', '5', '10', '19', '28']
        # self.content_layers = ['19']
        self.content_layers = ['21']
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.model.to(self.device)
        
        self.transform = transforms.Compose([
            transforms.Resize((224, 224)),
            transforms.ToTensor(),
            transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
        ])
        
        self.mse_loss = nn.MSELoss()

    def extract_features(self, image):
        image = self.transform(image).unsqueeze(0).to(self.device)
        features = {}
        x = image
        for name, layer in self.model._modules.items():
            x = layer(x)
            if name in self.style_layers or name in self.content_layers:
                features[name] = x
        return features

    def gram_matrix(self, features):
        B, C, H, W = features.size()
        F = features.view(B, C, H * W)
        G = torch.bmm(F, F.transpose(1, 2))
        return G
        # return G.div(H * W)


    def calculate_fid(self, mean1, cov1, mean2, cov2):
        diff = np.sum((mean1 - mean2) ** 2)
        cov_mean = sqrtm(np.dot(cov1, cov2))
        if np.iscomplexobj(cov_mean):
            cov_mean = cov_mean.real  # 去掉複數部分
        fid = diff + np.trace(cov1 + cov2 - 2 * cov_mean)
        return fid

    def compute_fid(self, result_features, style_features):
        def calculate_statistics(features):
            feat = features['5'].detach().cpu().numpy()
            feat = feat.reshape(feat.shape[0], feat.shape[1], -1)
            feat = np.mean(feat, axis=2)  # [B, C]

            # 更穩定的標準化
            # eps = 1e-10
            # feat_std = np.maximum(np.std(feat), eps)
            # feat = (feat - np.mean(feat)) / feat_std

            mu = np.mean(feat, axis=0)
            # 更穩定的協方差計算
            feat_centered = feat - mu
            sigma = np.dot(feat_centered.T, feat_centered) / max(feat.shape[0] - 1, 1)
            # sigma = sigma + np.eye(sigma.shape[0]) * eps

            return mu, sigma.astype(np.float64)  # 確保數值類型

        mu1, sigma1 = calculate_statistics(result_features)
        mu2, sigma2 = calculate_statistics(style_features)
        
        fid = self.calculate_fid(mu1, sigma1, mu2, sigma2)
        
        return fid

        # 計算 FID
        # diff = mu1 - mu2
        # try:
        #     covmean = sqrtm(sigma1.dot(sigma2))
        #     if np.iscomplexobj(covmean):
        #         covmean = covmean.real
        #     tr_covmean = np.trace(covmean)
        # except ValueError:
        #     # 如果計算失敗，返回大值
        #     return 1000.0

        # return float(diff.dot(diff) + np.trace(sigma1) + np.trace(sigma2) - 2 * tr_covmean)

    


    def compute_metrics(self, result_features_gray, result_features, style_features, content_features):
        style_weights = {'0': 1.0, '5': 0.75, '10': 0.2, '19': 0.2, '28': 0.2}
        style_loss = 0.0
        style_sim = 0.0
        for layer in self.style_layers:
            B, C, H, W = result_features_gray[layer].size()
            result_gram = self.gram_matrix(result_features_gray[layer])
            style_gram = self.gram_matrix(style_features[layer])
            
            # Loss
            style_loss += style_weights[layer] * torch.sum((style_gram - result_gram) ** 2) / (4 * (C ** 2) * (H * W) ** 2)

            # Similarity
            sim = torch.dot(
                    result_features_gray[layer].view(-1), 
                    style_features[layer].view(-1)
                ) / (
                    torch.norm(result_features_gray[layer].view(-1)) * 
                    torch.norm(style_features[layer].view(-1))
                )
            style_sim += style_weights[layer] * sim.item()
        
        content_loss = 0.0
        content_sim = 0.0
        for layer in self.content_layers:
            result_feat = result_features[layer]
            content_feat = content_features[layer]
            
            # Loss
            content_loss += self.mse_loss(result_feat, content_feat)
            
            # Similarity
            result_flat = result_feat.view(-1)
            content_flat = content_feat.view(-1)
            sim = torch.dot(result_flat, content_flat) / (torch.norm(result_flat) * torch.norm(content_flat))
            content_sim += sim.item()
        
        return {
            'style_loss': style_loss.item() / sum(style_weights.values()),
            'style_similarity': style_sim / sum(style_weights.values()),
            'content_loss': content_loss.item(),
            'content_similarity': content_sim / len(self.content_layers),
            'fid': self.compute_fid(result_features, style_features)
        }

    def evaluate_group(self, group_name):
        group_dir = os.path.join(self.root_dir, group_name)
        group_id = group_name[-1]
        results = {method: [] for method in self.methods}

        for pair_name in sorted(os.listdir(group_dir)):
            print("pair_name", pair_name)
            style_id, content_id = pair_name.split("_to_")
            style_path = os.path.join(self.style_dir, f"style{group_id}", f"{style_id[-2:]}.jpg")
            content_path = os.path.join(self.content_dir, f"content{group_id}", f"{content_id[-2:]}.jpg")
            
            if not os.path.exists(style_path) or not os.path.exists(content_path):
                continue

            style_image = Image.open(style_path).convert("RGB")
            style_image = style_image.convert('L').convert('RGB')  # 轉灰階再轉RGB
            content_image = Image.open(content_path).convert("RGB")
            style_features = self.extract_features(style_image)
            content_features = self.extract_features(content_image)

            for method in self.methods:
                pair_path = os.path.join(group_dir, pair_name, method)
                if not os.path.isdir(pair_path):
                    continue

                pair_metrics = []
                for result_name in sorted(os.listdir(pair_path)):
                    result_path = os.path.join(pair_path, result_name)
                    result_image = Image.open(result_path).convert("RGB")
                    result_image_gray = result_image.convert('L').convert('RGB')  # 轉灰階再轉RGB
                    if self.is_black_image(result_image):
                        print("black image")
                        continue
                        
                    result_features = self.extract_features(result_image)
                    result_features_gray = self.extract_features(result_image_gray)
                    metrics = self.compute_metrics(result_features_gray, result_features, style_features, content_features)
                    pair_metrics.append(metrics)

                if pair_metrics:
                    avg_metrics = {k: np.mean([m[k] for m in pair_metrics]) for k in pair_metrics[0]}
                    results[method].append(avg_metrics)

        return results

    def evaluate_all_groups(self, groups, output_path="evaluation_results.csv"):
        all_results = []
        raw_metrics = {
            'style_loss': [],
            'style_similarity': [],
            'content_loss': [],
            'content_similarity': [],
            'fid': []
        }
        
        for method in self.methods:
            method_metrics = {
                'style_loss': [], 'style_similarity': [],
                'content_loss': [], 'content_similarity': [],
                'fid': []
            }
            
            for group in groups:
                print(f"Processing {group} - {method}")
                group_results = self.evaluate_group(group)
                if method in group_results and group_results[method]:
                    for metric in method_metrics:
                        values = [r[metric] for r in group_results[method]]
                        method_metrics[metric].extend(values)
                        
                # 存儲平均值用於比較
            for metric in raw_metrics:
                raw_metrics[metric].append(np.mean(method_metrics[metric]))
            
            result = {
                'method': method,
                'style_loss': f"{np.mean(method_metrics['style_loss']):.3f} ± {np.std(method_metrics['style_loss']):.3f}",
                'style_similarity': f"{np.mean(method_metrics['style_similarity']):.3f} ± {np.std(method_metrics['style_similarity']):.3f}",
                'content_loss': f"{np.mean(method_metrics['content_loss']):.3f} ± {np.std(method_metrics['content_loss']):.3f}",
                'content_similarity': f"{np.mean(method_metrics['content_similarity']):.3f} ± {np.std(method_metrics['content_similarity']):.3f}",
                'fid': f"{np.mean(method_metrics['fid']):.3f} ± {np.std(method_metrics['fid']):.3f}"
            }
            all_results.append(result)
            print(f"\nMethod: {method}")
            print(f"Style Loss: {result['style_loss']}")
            print(f"Style Similarity: {result['style_similarity']}")
            print(f"Content Loss: {result['content_loss']}")
            print(f"Content Similarity: {result['content_similarity']}\n")
            print(f"FID: {result['fid']}\n")

        # 創建DataFrame
        df = pd.DataFrame(all_results)
        df.to_excel(output_path, index=False)

        # 載入workbook進行顏色標記
        wb = load_workbook(output_path)
        ws = wb.active
        # 定義顏色
        highlight_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # 指標和它們的最優方向
        metrics = {
            'style_loss': (2, 'min'),
            'style_similarity': (3, 'max'),
            'content_loss': (4, 'min'),
            'content_similarity': (5, 'max'),
            'fid': (6, 'min')
        }
        
        # 標記最佳值
        for metric, (col, direction) in metrics.items():
            values = raw_metrics[metric]
            best_idx = np.argmin(values) if direction == 'min' else np.argmax(values)
            cell = ws[f"{get_column_letter(col)}{best_idx + 2}"]
            cell.fill = highlight_fill

        wb.save(output_path)
            
    @staticmethod
    def is_black_image(image):
        return np.all(np.array(image) <= 10)

if __name__ == "__main__":
    root_dir = "output"
    style_dir = "data/Style"
    content_dir = "data/Content"
    # groups = ["group3"]
    groups = ["group1", "group2", "group3", "group4", "group5"]
    # methods = ["original"]
    methods = ["original", "lab_transfer", "luv_transfer", "pca_transfer"]

    evaluator = StyleTransferEvaluator(root_dir, style_dir, content_dir, methods)
    evaluator.evaluate_all_groups(groups)